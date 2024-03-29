#!/usr/bin/env python3
"""

pip3 install garth requests readchar


export EMAIL=<your garmin email>

export PASSWORD=<your garmin password>

"""
import sys
import datetime

import pytz

import json
import logging
import os
import sys
from getpass import getpass

import openpyxl


import readchar

import requests

from garth.exc import GarthHTTPError

from garminconnect import (

    Garmin,

    GarminConnectAuthenticationError,

    GarminConnectConnectionError,

    GarminConnectTooManyRequestsError,
)


# Configure debug logging

# logging.basicConfig(level=logging.DEBUG)

logging.basicConfig(level=logging.INFO)

logger = logging.getLogger(__name__)


# Load environment variables if defined

email = os.getenv("EMAIL")

password = os.getenv("PASSWORD")

tokenstore = os.getenv("GARMINTOKENS") or "~/.garminconnect"

tokenstore_base64 = os.getenv("GARMINTOKENS_BASE64") or "~/.garminconnect_base64"

api = None


# Example selections and settings

today = datetime.date.today()

startdate = today - datetime.timedelta(days=7)  # Select past week

yesterday = today - datetime.timedelta(days=1)


def get_credentials():

    """Get user credentials."""


    email = input("Login e-mail: ")

    password = getpass("Enter password: ")


    return email, password



def init_api(email, password):

    """Initialize Garmin API with your credentials."""


    try:

        # Using Oauth1 and OAuth2 token files from directory
        print(

            f"Trying to login to Garmin Connect using token data from directory '{tokenstore}'...\n"
        )


        # Using Oauth1 and Oauth2 tokens from base64 encoded string

        # print(

        #     f"Trying to login to Garmin Connect using token data from file '{tokenstore_base64}'...\n"

        # )

        # dir_path = os.path.expanduser(tokenstore_base64)

        # with open(dir_path, "r") as token_file:

        #     tokenstore = token_file.read()


        garmin = Garmin()

        garmin.login(tokenstore)


    except (FileNotFoundError, GarthHTTPError, GarminConnectAuthenticationError):

        # Session is expired. You'll need to log in again
        print(

            "Login tokens not present, login with your Garmin Connect credentials to generate them.\n"

            f"They will be stored in '{tokenstore}' for future use.\n"
        )

        try:

            # Ask for credentials if not set as environment variables

            if not email or not password:

                email, password = get_credentials()


            garmin = Garmin(email, password)

            garmin.login()

            # Save Oauth1 and Oauth2 token files to directory for next login

            garmin.garth.dump(tokenstore)
            print(

                f"Oauth tokens stored in '{tokenstore}' directory for future use. (first method)\n"
            )

            # Encode Oauth1 and Oauth2 tokens to base64 string and safe to file for next login (alternative way)

            token_base64 = garmin.garth.dumps()

            dir_path = os.path.expanduser(tokenstore_base64)

            with open(dir_path, "w") as token_file:

                token_file.write(token_base64)
            print(

                f"Oauth tokens encoded as base64 string and saved to '{dir_path}' file for future use. (second method)\n"
            )

        except (FileNotFoundError, GarthHTTPError, GarminConnectAuthenticationError, requests.exceptions.HTTPError) as err:

            logger.error(err)

            return None

    return garmin


def convert_gmt_timestamp_to_local(gmt_timestamp, local_timezone='America/Los_Angeles'):

    # Convert milliseconds to seconds

    timestamp_seconds = gmt_timestamp / 1000


    # Create a timezone-aware UTC datetime object

    utc_datetime = datetime.datetime.fromtimestamp(timestamp_seconds, tz=datetime.timezone.utc)


    # Convert to local time (Pacific Time)

    local_timezone_obj = pytz.timezone(local_timezone)

    local_datetime = utc_datetime.astimezone(local_timezone_obj)


    return local_datetime.replace(tzinfo=None)


def user_input(message, valueType, min, max):

    while True:

        try:

            if valueType == 'int':

                userInput = int(input(message + " "))

            elif valueType == 'bool':

                userInput = input(message + " ")

                if userInput.lower() not in ["y", "n"]:

                    print("Error: please enter y or n")
                    continue

                elif userInput.lower() == "y":

                    return True

                else:

                    return False

            elif valueType == 'float':

                userInput = float(input(message + " "))

            if min <= userInput <= max:

                return userInput

            else:

                print("Invalid input. Try again.")

        except ValueError:

            print(f"Invalid input. Please enter a valid {valueType}")


api = init_api(email, password)

if api == False:

    print("failed to login and initialize api")

    sys.exit()


#getting yesterday's relevent Health Data

releventHealthData = {"totalKilocalories": 0, "totalSteps": 0, "highlyActiveSeconds": 0, 

    "activeSeconds": 0, "sedentarySeconds": 0, "bodyBatteryDrainedValue": 0, "bodyBatteryChargedValue": 0, 

    "bodyBatteryLowestValue": 0, "bodyBatteryHighestValue": 0, "avgWakingRespirationValue": 0 

    }

stats = api.get_stats(yesterday.isoformat())

for key in releventHealthData:

    if key in stats:

        releventHealthData[key] = stats[key]

print(f"releventHealthData: {releventHealthData}")
print()


## gathering yesterday's relevent heart data

releventHeartrateData = {'maxHeartRate': 0, 'restingHeartRate': 0}

heartData = api.get_heart_rates(yesterday.isoformat())

for key in releventHeartrateData:

    if key in heartData:

        releventHeartrateData[key] = heartData[key]

print(f"releventHeartrateData: {releventHeartrateData}")
print()


## gathering yesterday's relevent stress data

releventStressData = {'maxStressLevel': 0, 'avgStressLevel': 0}

stressData = api.get_stress_data(yesterday.isoformat())

for key in releventStressData:

    if key in stressData:

        releventStressData[key] = stressData[key]

print(f"releventStressData: {releventStressData}")
print()


## gathering last night's relevent sleep data

releventSleepData = {'sleepWindowConfirmed': False, 'sleepTimeHours': 0, 'sleepStartTime': 0, 'sleepEndTime': 0, 'sleepScore': 0}

sleepData = api.get_sleep_data(today)

sleepSummary = sleepData['dailySleepDTO']

releventSleepData['sleepWindowConfirmed'] = sleepData['dailySleepDTO']['sleepWindowConfirmed']

if releventSleepData['sleepWindowConfirmed'] == True: 

    releventSleepData['sleepTimeHours'] = sleepSummary['sleepTimeSeconds'] / 3600

    sleepStartTime = sleepSummary['sleepStartTimestampGMT']

    releventSleepData['sleepStartTime'] = convert_gmt_timestamp_to_local(sleepStartTime)

    sleepEndTime = sleepSummary['sleepEndTimestampGMT']

    releventSleepData['sleepEndTime'] = convert_gmt_timestamp_to_local(sleepEndTime)

    releventSleepData['sleepScore'] = sleepSummary['sleepScores']['overall']['value']

    print(f"releventSleepData: {releventSleepData}")
    print()

else:

    print("No sleep data available from last night")

    sys.exit()


## gathering user input stats
userData = {'wakefulness': 0, 'cupsOfCoffee': 0, 'modafanil': 0, 'exerciseYesterday': False, 'exerciseToday': False, 'focus': 0, 'dayScore': 0, 'enjoyability': 0}
userData['wakefulness'] = user_input("How awake did you feel today? (1-10)", 'int', 1, 10)
userData['cupsOfCoffee'] = user_input("How many cups of coffee did you have today? (0-10)", 'int', 0, 10)
userData['modafanil'] = user_input("How much modafanil did you have today? (0-2 in steps of 0.25)", 'float', 0, 2)
userData['exerciseYesterday'] = user_input("Did you exercise yesterday? (y/n)", 'bool', False, True)
userData['exerciseToday'] = user_input("Did you exercise today? (y/n)", 'bool', False, True)
userData['focus'] = user_input("How well were you able to focus today? (1-10)", 'int', 1, 10)
userData['dayScore'] = user_input("On a scale from 1-10, was today a bad day (1) or a good day (10)?", 'int', 1, 10)
userData['enjoyability'] = user_input("How enjoyable were the tasks you worked on today?", 'int', 1, 10)
print() 


## putting data into one dictionary

allData = {"date": today, **userData, **releventSleepData, **releventHealthData, **releventHeartrateData, **releventStressData}

#print(allData)

#print()

numStats = len(allData)


print(f"Number of tracked stats: {numStats}")




## writing data to excel file

# opening excel file

excelWorkbook = openpyxl.load_workbook('sleepdata.xlsx')


# selecting sheet

sheet_by_name = excelWorkbook['data']

sheet = excelWorkbook.active


# reading data names from excel

colNames = []

for row in sheet.iter_rows(min_row=1, max_col=numStats, max_row=1, values_only=True):

    for value in row:

        colNames.append(value)


# if excel sheet is empty, load in variable names

dataNames = list(allData.keys())

if colNames[0] == None:

    index = 0

    for row in sheet.iter_rows(min_row=1, max_col=numStats, max_row=1):

        for cell in row:

            cell.value = dataNames[index]

            index = index + 1


# finding first new line

newRow = 1

while(newRow < 10):

    cell = sheet.cell(row = newRow, column = 1)

    if cell.value == None:

        break

    else:

        newRow = newRow + 1


# checking to see if data has been logged today

if newRow > 2 and sheet.cell(row = newRow-1, column = 1).value.date() == today:

    reenter = user_input("Data already entered for today. Would you like to re-enter data? (y/n)", 'bool', 0, 1)

    if reenter:

        newRow = newRow - 1

    else: 

        sys.exit()


# writing data to sheet

index = 0

for row in sheet.iter_rows(min_row=newRow, max_col=numStats, max_row=newRow):

    for cell in row:

        cell.value = allData[dataNames[index]]

        index = index + 1


excelWorkbook.save('sleepdata.xlsx')

print("data saved.")